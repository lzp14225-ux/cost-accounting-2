"""
NCTimeAgent - NC时间计算Agent
负责人：人员B1

职责：
1. 调用外部 NC Agent 获取每个子图的 NC 时间数据
2. 解析 NC 返回的 JSON 数据
3. 将时间数据写入 subgraphs 表（nc_roughing_time, nc_milling_time, drilling_time）
4. 将详细时间数据写入 features 表（nc_time_cost 字段）
"""
from typing import Dict, Any, List, Optional
import asyncio
import httpx
import re
import os
import json
import time
import logging
from decimal import Decimal
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from dotenv import load_dotenv
from tenacity import retry, stop_after_attempt, wait_exponential
from sqlalchemy import select, update
from .base_agent import BaseAgent
from shared.database import get_db
from shared.models import Job, Subgraph, Feature

load_dotenv()
module_logger = logging.getLogger(__name__)

class NCTimeAgent(BaseAgent):
    """
    NC时间计算Agent
    调用外部统一NC Agent计算钻孔、开粗、精铣时间
    """

    EXCEL_FACE_CODE_MAP = {
        "A面": "Z",
        "B面": "B",
        "C面": "C",
        "D面": "C_B",
        "E面": "Z_VIEW",
        "F面": "B_VIEW",
    }

    JSON_FACE_CODE_MAP = {
        "A": "Z",
        "B": "B",
        "C": "C",
        "D": "C_B",
        "E": "Z_VIEW",
        "F": "B_VIEW",
    }
    
    def __init__(self, nc_agent_url: str = None, progress_publisher=None):
        super().__init__("NCTimeAgent")
        self.nc_agent_url = nc_agent_url or os.environ["NC_AGENT_URL"]
        self.timeout = int(os.environ["NC_AGENT_TIMEOUT"])
        self.request_timeout = int(os.environ["NC_AGENT_REQUEST_TIMEOUT"])
        self.poll_interval = float(os.environ["NC_AGENT_POLL_INTERVAL"])
        self.progress_publisher = progress_publisher
        self.logs_root = Path("logs")
        self.nc_log_retention_days = int(os.getenv("NC_LOG_RETENTION_DAYS", "7"))
        self.nc_excel_retention_days = int(os.getenv("NC_EXCEL_RETENTION_DAYS", "7"))
        self.nc_log_cleanup_interval_days = int(os.getenv("NC_LOG_CLEANUP_INTERVAL_DAYS", "7"))
        self.nc_log_cleanup_state_file = self.logs_root / ".nc_log_cleanup_state.json"
        self.nc_excel_dir = os.getenv("NC_EXCEL_DIR", "").strip()
        self.nc_excel_logs_dir = self.logs_root / "ncexcel"
        self.logger.info(
            f"[NCTimeAgent] initialized: url={self.nc_agent_url}, "
            f"workflow_timeout={self.timeout}s, request_timeout={self.request_timeout}s, "
            f"poll_interval={self.poll_interval}s"
        )

    def _log_visible(self, level: str, message: str):
        log_method = getattr(self.logger, level, self.logger.info)
        module_log_method = getattr(module_logger, level, module_logger.info)
        log_method(message)
        module_log_method(message)

    async def process(self, context: Dict[str, Any]) -> Dict[str, Any]:
        """
        处理NC时间计算
        
        Args:
            context: {
                "job_id": "任务ID",
                "dwg_file_path": "DWG文件路径（MinIO路径或本地路径）",
                "prt_file_path": "PRT文件路径（MinIO路径或本地路径）"
            }
        
        Returns:
            {
                "status": "ok" | "error",
                "message": "消息",
                "summary": {
                    "total_subgraphs": 总子图数,
                    "success_count": 成功数,
                    "failed_count": 失败数
                }
            }
        """
        job_id = context.get("job_id")
        if not job_id:
            return {"status": "error", "message": "缺少 job_id 参数"}
        
        self.logger.info(f"[NCTimeAgent] 开始处理 NC 时间计算: job_id={job_id}")

        # Excel 解析链路已停用，保留代码作参考，不删除。
        # excel_result = await self._process_nc_excel_workbooks(job_id, context)
        # if excel_result is not None:
        #     return excel_result

        minio_json_result = await self._process_nc_json_from_minio(job_id)
        if minio_json_result is not None:
            return minio_json_result

        # 旧逻辑保留作参考：此前 MinIO JSON 未命中时，会继续回退到外部 NC HTTP 服务。
        # 现在这条链路已停用，只保留最新的 MinIO JSON 解析逻辑。
        #
        # # 发布进度：NC 时间计算开始
        # self._maybe_cleanup_nc_log_files()
        # self._maybe_cleanup_nc_log_files()
        # if self.progress_publisher:
        #     from shared.progress_stages import ProgressStage, ProgressPercent
        #     self.progress_publisher.publish_progress(
        #         job_id=job_id,
        #         stage=ProgressStage.NC_CALCULATION_STARTED,
        #         progress=ProgressPercent.NC_CALCULATION_STARTED,
        #         message="NC 时间计算开始",
        #         details={"nc_agent_url": self.nc_agent_url}
        #     )
        #
        # # 临时文件列表（用于清理）
        # temp_files = []
        #
        # try:
        #     # 1. 获取文件路径（如果是 MinIO 路径，需要先下载）
        #     dwg_file_path = context.get("dwg_file_path")
        #     prt_file_path = context.get("prt_file_path")
        #
        #     if not dwg_file_path or not prt_file_path:
        #         return {"status": "error", "message": "缺少 dwg_file_path 或 prt_file_path 参数"}
        #
        #     # 检查是否为 MinIO 路径（不以 / 或盘符开头）
        #     dwg_local_path = await self._ensure_local_file(dwg_file_path, job_id, "dwg")
        #     prt_local_path = await self._ensure_local_file(prt_file_path, job_id, "prt")
        #
        #     if dwg_local_path != dwg_file_path:
        #         temp_files.append(dwg_local_path)
        #     if prt_local_path != prt_file_path:
        #         temp_files.append(prt_local_path)
        #
        #     # 2. 调用外部 NC Agent
        #     nc_result = await self.call_nc_agent_workflow(
        #         job_id=job_id,
        #         dwg_file=dwg_local_path,
        #         prt_file=prt_local_path
        #     )
        #
        #     fail_itemcodes = self._extract_fail_itemcodes(nc_result)
        #     await self._save_nc_failed_itemcodes(job_id, fail_itemcodes)
        #
        #     return await self._process_nc_result_payload(job_id, nc_result, source="http")
        #
        # except Exception as e:
        #     self.logger.error(f"[NCTimeAgent] NC 时间计算失败: {e}", exc_info=True)
        #
        #     # 发布进度：NC 时间计算失败
        #     if self.progress_publisher:
        #         from shared.progress_stages import ProgressStage, ProgressPercent
        #         self.progress_publisher.publish_progress(
        #             job_id=job_id,
        #             stage=ProgressStage.NC_CALCULATION_FAILED,
        #             progress=ProgressPercent.NC_CALCULATION_STARTED,
        #             message=f"NC 时间计算失败: {str(e)}",
        #             details={"error": str(e)}
        #         )
        #
        #     return {"status": "error", "message": f"NC 时间计算失败: {str(e)}"}
        #
        # finally:
        #     # 清理临时文件
        #     await self._cleanup_temp_files(temp_files)

        error_message = "NC 时间计算失败: 未命中可用的 MinIO JSON 数据，且外部 NC HTTP 回退链路已停用"
        self.logger.error(f"[NCTimeAgent] {error_message}")
        if self.progress_publisher:
            from shared.progress_stages import ProgressStage, ProgressPercent
            self.progress_publisher.publish_progress(
                job_id=job_id,
                stage=ProgressStage.NC_CALCULATION_FAILED,
                progress=ProgressPercent.NC_CALCULATION_STARTED,
                message=error_message,
                details={"error": error_message, "source": "minio_json_only"}
            )
        return {"status": "error", "message": error_message}
    
    async def call_nc_agent_workflow(
        self,
        job_id: str,
        dwg_file: str = None,
        prt_file: str = None
    ) -> Dict[str, Any]:
        """Call NC Agent using submit/status/result workflow API."""
        self.logger.info(f"[NCTimeAgent] calling NC workflow API: {self.nc_agent_url}")

        if not prt_file or not dwg_file:
            raise ValueError("Missing required nc input files: prt_file/dwg_file")

        files: Dict[str, Any] = {}
        try:
            dwg_ext = os.path.splitext(dwg_file)[1].lower()
            prt_ext = os.path.splitext(prt_file)[1].lower()

            files["prt_file"] = (f"model{prt_ext}", open(prt_file, "rb"), "application/octet-stream")
            drawing_field = "dxf_file" if dwg_ext == ".dxf" else "dwg_file"
            files[drawing_field] = (f"drawing{dwg_ext}", open(dwg_file, "rb"), "application/octet-stream")

            data = {
                "skip_approval": "true",
                "auto_continue": "true"
            }

            self.logger.info(
                f"[NCTimeAgent] submit files: prt={prt_file}, {drawing_field}={dwg_file}, "
                f"workflow_timeout={self.timeout}s, request_timeout={self.request_timeout}s"
            )

            async with httpx.AsyncClient(timeout=self.request_timeout) as client:
                submit_payload = await self._submit_nc_workflow(client, files, data)
                task_id = submit_payload.get("task_id")
                if not task_id:
                    raise RuntimeError(f"NC submit missing task_id: {submit_payload}")

                self.logger.info(
                    f"[NCTimeAgent] NC task submitted: task_id={task_id}, "
                    f"status={submit_payload.get('status')}, trace_id={submit_payload.get('trace_id')}"
                )

                await self._wait_for_nc_workflow(client, task_id)
                result = await self._fetch_nc_result(client, task_id)
                self._save_nc_debug_response(job_id, result)
                return result
        finally:
            for file_tuple in files.values():
                if len(file_tuple) > 1 and hasattr(file_tuple[1], "close"):
                    file_tuple[1].close()

    async def _submit_nc_workflow(
        self,
        client: httpx.AsyncClient,
        files: Dict[str, Any],
        data: Dict[str, Any]
    ) -> Dict[str, Any]:
        response = await client.post(
            f"{self.nc_agent_url}/api/v1/workflow/3d/submit",
            files=files,
            data=data
        )
        response.raise_for_status()
        payload = response.json()
        if not payload.get("success"):
            raise RuntimeError(f"NC submit failed: {payload.get('message') or payload}")
        return payload

    async def _wait_for_nc_workflow(self, client: httpx.AsyncClient, task_id: str):
        deadline = time.monotonic() + self.timeout
        last_status = None
        last_message = ""

        while time.monotonic() < deadline:
            response = await client.get(f"{self.nc_agent_url}/api/v1/workflow/3d/status/{task_id}")
            response.raise_for_status()
            payload = response.json()

            status = payload.get("status")
            message = payload.get("message", "")
            if status != last_status:
                self.logger.info(
                    f"[NCTimeAgent] NC status: task_id={task_id}, status={status}, message={message}"
                )
                last_status = status
            last_message = message

            if status == "success":
                return

            if status == "failed":
                raise RuntimeError(f"NC workflow failed: {self._extract_nc_error_message(payload)}")

            if status not in {"queued", "running"}:
                raise RuntimeError(f"NC workflow returned unknown status: {status}, payload={payload}")

            await asyncio.sleep(self.poll_interval)

        raise RuntimeError(
            f"NC workflow timeout: task_id={task_id}, timeout={self.timeout}s, "
            f"last_status={last_status}, last_message={last_message}"
        )

    async def _fetch_nc_result(self, client: httpx.AsyncClient, task_id: str) -> Dict[str, Any]:
        response = await client.get(f"{self.nc_agent_url}/api/v1/workflow/3d/result/{task_id}")
        response.raise_for_status()
        payload = response.json()

        status = payload.get("status")
        if status == "failed" or payload.get("success") is False:
            raise RuntimeError(f"NC result failed: {self._extract_nc_error_message(payload)}")

        result = payload.get("result")
        if not isinstance(result, dict):
            raise RuntimeError(f"NC result payload invalid: {payload}")

        self.logger.info(
            f"[NCTimeAgent] NC result fetched: task_id={task_id}, status={status}, "
            f"code={result.get('code')}, message={result.get('message')}"
        )
        return result

    def _extract_nc_error_message(self, payload: Dict[str, Any]) -> str:
        error = payload.get("error")
        if isinstance(error, dict):
            return error.get("error_message") or error.get("message") or payload.get("message") or str(payload)
        return payload.get("message") or str(error) or str(payload)

    def _save_nc_debug_response(self, job_id: str, result: Dict[str, Any]):
        debug_dir = Path("logs") / "nc_agent_debug"
        debug_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        debug_file = debug_dir / f"response_{job_id}_{timestamp}.json"
        with open(debug_file, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        self.logger.info(f"[NCTimeAgent] debug response saved: {debug_file}")

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=2, max=10)
    )
    async def call_nc_agent(
        self,
        job_id: str,
        dwg_file: str = None,
        prt_file: str = None
    ) -> Dict[str, Any]:
        """
        调用外部 NC Agent (NC 3D Workflow API)
        
        Args:
            job_id: 任务ID
            dwg_file: DWG文件路径（保持原始格式）
            prt_file: PRT文件路径
        
        Returns:
            NC Agent 返回的 JSON 数据
        """
        self.logger.info(f"[NCTimeAgent] 调用 NC Agent: {self.nc_agent_url}")
        
        if not prt_file or not dwg_file:
            raise ValueError("必须提供 prt_file 和 dwg_file 路径")
        
        # 准备文件上传
        files = {}
        try:
            # 获取原始文件扩展名
            import os
            dwg_ext = os.path.splitext(dwg_file)[1]  # .dwg 或 .dxf
            prt_ext = os.path.splitext(prt_file)[1]  # .prt
            
            # 使用原始文件名和扩展名
            files['prt_file'] = (f'model{prt_ext}', open(prt_file, 'rb'), 'application/octet-stream')
            files['dxf_file'] = (f'drawing{dwg_ext}', open(dwg_file, 'rb'), 'application/octet-stream')
            
            # 准备表单数据
            data = {
                'skip_approval': 'true',
                'auto_continue': 'true'
            }
            
            self.logger.info(f"[NCTimeAgent] 上传文件并等待处理完成: prt={prt_file}, dwg={dwg_file}")
            self.logger.info(f"[NCTimeAgent] 超时设置: {self.timeout}秒")
            
            async with httpx.AsyncClient(timeout=self.timeout) as client:
                response = await client.post(
                    f"{self.nc_agent_url}/api/v1/workflow/3d/run",
                    files=files,
                    data=data
                )
                response.raise_for_status()
                result = response.json()
                
                # 记录返回的 code
                code = result.get("code")
                message = result.get("message", "")
                self.logger.info(f"[NCTimeAgent] NC Agent 返回: code={code}, message={message}")
                
                # 详细记录响应结构（用于调试）
                import json
                data_obj = result.get("data", {})
                self.logger.info(f"[NCTimeAgent] 响应结构详情:")
                self.logger.info(f"  - data 存在: {data_obj is not None}")
                self.logger.info(f"  - data 类型: {type(data_obj)}")
                if data_obj:
                    self.logger.info(f"  - task_id: {data_obj.get('task_id')}")
                    self.logger.info(f"  - execution_time: {data_obj.get('execution_time')}")
                    self.logger.info(f"  - output_dir: {data_obj.get('output_dir')}")
                    self.logger.info(f"  - steps_completed: {data_obj.get('steps_completed')}")
                    self.logger.info(f"  - failed_at_step: {data_obj.get('failed_at_step')}")
                    
                    json_output = data_obj.get("json_output")
                    self.logger.info(f"  - json_output 存在: {json_output is not None}")
                    self.logger.info(f"  - json_output 类型: {type(json_output)}")
                    if json_output and isinstance(json_output, dict):
                        self.logger.info(f"  - json_output 子图数量: {len(json_output)}")
                        self.logger.info(f"  - json_output 子图列表: {list(json_output.keys())[:5]}")  # 只显示前5个
                    
                    # 如果有错误详情，记录下来
                    error_details = data_obj.get("error_details")
                    if error_details:
                        self.logger.warning(f"  - error_details: {error_details[:500]}")
                
                # 保存完整响应到文件（用于调试）
                try:
                    from pathlib import Path
                    from datetime import datetime
                    debug_dir = Path("logs") / "nc_agent_debug"
                    debug_dir.mkdir(parents=True, exist_ok=True)
                    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                    debug_file = debug_dir / f"response_{job_id}_{timestamp}.json"
                    with open(debug_file, 'w', encoding='utf-8') as f:
                        json.dump(result, f, ensure_ascii=False, indent=2)
                    self.logger.info(f"[NCTimeAgent] 完整响应已保存: {debug_file}")
                except Exception as e:
                    self.logger.warning(f"[NCTimeAgent] 保存响应文件失败: {e}")
                
                return result
                
        finally:
            # 关闭文件句柄
            for file_tuple in files.values():
                if len(file_tuple) > 1 and hasattr(file_tuple[1], 'close'):
                    file_tuple[1].close()

    def _maybe_cleanup_nc_log_files(self):
        """Clean NC log json files on a fixed interval."""
        if (
            self.nc_log_retention_days <= 0
            and self.nc_excel_retention_days <= 0
        ) or self.nc_log_cleanup_interval_days <= 0:
            self.logger.info(
                "[NCTimeAgent] Skip NC log cleanup: "
                f"retention_days={self.nc_log_retention_days}, "
                f"excel_retention_days={self.nc_excel_retention_days}, "
                f"cleanup_interval_days={self.nc_log_cleanup_interval_days}"
            )
            return

        now_ts = time.time()
        interval_seconds = self.nc_log_cleanup_interval_days * 24 * 60 * 60

        try:
            last_cleanup_ts = self._load_nc_log_cleanup_state()
            if last_cleanup_ts and now_ts - last_cleanup_ts < interval_seconds:
                return

            deleted_files = 0
            if self.nc_log_retention_days > 0:
                cutoff_ts = now_ts - (self.nc_log_retention_days * 24 * 60 * 60)
                deleted_files += self._cleanup_json_files(self.logs_root / "nc_agent_debug", cutoff_ts)
                deleted_files += self._cleanup_json_files(self.logs_root / "nc_responses", cutoff_ts)
            if self.nc_excel_retention_days > 0:
                excel_cutoff_ts = now_ts - (self.nc_excel_retention_days * 24 * 60 * 60)
                deleted_files += self._cleanup_excel_files(self.nc_excel_logs_dir, excel_cutoff_ts)
            self._write_nc_log_cleanup_state(now_ts)
            self.logger.info(
                "[NCTimeAgent] NC log cleanup completed: "
                f"deleted_files={deleted_files}, retention_days={self.nc_log_retention_days}, "
                f"excel_retention_days={self.nc_excel_retention_days}, "
                f"cleanup_interval_days={self.nc_log_cleanup_interval_days}"
            )
        except Exception as e:
            self.logger.warning(f"[NCTimeAgent] NC log cleanup failed: {e}")

    def _load_nc_log_cleanup_state(self) -> float:
        if not self.nc_log_cleanup_state_file.exists():
            return 0.0

        try:
            state = json.loads(self.nc_log_cleanup_state_file.read_text(encoding="utf-8"))
            return float(state.get("last_cleanup_ts", 0.0))
        except Exception as e:
            self.logger.warning(f"[NCTimeAgent] Failed to read cleanup state: {e}")
            return 0.0

    def _write_nc_log_cleanup_state(self, timestamp: float):
        self.logs_root.mkdir(parents=True, exist_ok=True)
        payload = {
            "last_cleanup_ts": timestamp,
            "last_cleanup_at": datetime.utcfromtimestamp(timestamp).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "retention_days": self.nc_log_retention_days,
            "cleanup_interval_days": self.nc_log_cleanup_interval_days,
        }
        self.nc_log_cleanup_state_file.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def _cleanup_json_files(self, root_dir: Path, cutoff_ts: float) -> int:
        return self._cleanup_files_by_pattern(root_dir, cutoff_ts, "*.json")

    def _cleanup_excel_files(self, root_dir: Path, cutoff_ts: float) -> int:
        return self._cleanup_files_by_pattern(root_dir, cutoff_ts, "*.xlsx")

    def _cleanup_files_by_pattern(self, root_dir: Path, cutoff_ts: float, pattern: str) -> int:
        if not root_dir.exists():
            return 0

        deleted_files = 0
        for file_path in root_dir.rglob(pattern):
            try:
                if file_path.stat().st_mtime < cutoff_ts:
                    file_path.unlink()
                    deleted_files += 1
            except FileNotFoundError:
                continue
            except Exception as e:
                self.logger.warning(f"[NCTimeAgent] Failed to delete expired log file {file_path}: {e}")

        directories = [path for path in root_dir.rglob("*") if path.is_dir()]
        directories.sort(key=lambda path: len(path.parts), reverse=True)
        for directory in directories:
            try:
                if not any(directory.iterdir()):
                    directory.rmdir()
            except OSError:
                continue

        return deleted_files

    def _extract_subgraph_id(self, subgraph_name: str) -> str:
        """
        从子图名称中提取短ID
        
        例如：
        - PH-01-M250297-P5.json -> PH-01
        - PU-06-M250297-P5.json -> PU-06
        
        Args:
            subgraph_name: 子图名称
        
        Returns:
            短ID（如 PH-01）
        """
        # 移除 .json 后缀
        name = subgraph_name.replace(".json", "")
        
        # 提取前两段（例如：PH-01）
        parts = name.split("-")
        if len(parts) >= 2:
            return f"{parts[0]}-{parts[1]}"
        
        return name
    
    async def _find_subgraph_id(self, job_id: str, short_id: str) -> str:
        """
        根据短ID查找完整的 subgraph_id
        
        Args:
            job_id: 任务ID
            short_id: 短ID（如 PH-01）
        
        Returns:
            完整的 subgraph_id（如 2cd0b581-f2b9-481d-a2fd-33074f57ebd4_PH-01）
        """
        import uuid
        job_uuid = uuid.UUID(job_id)
        
        async for db in get_db():
            # 查询所有子图
            result = await db.execute(
                select(Subgraph.subgraph_id)
                .where(Subgraph.job_id == job_uuid)
            )
            subgraph_ids = [row[0] for row in result.fetchall()]
            
            # 查找匹配的子图ID（以 short_id 结尾）
            for subgraph_id in subgraph_ids:
                if subgraph_id.endswith(f"_{short_id}") or subgraph_id == short_id:
                    return subgraph_id
            
            break
        
        return None
    
    async def _ensure_local_file(self, file_path: str, job_id: str, file_type: str) -> str:
        """
        确保文件在本地文件系统中
        
        如果是 MinIO 路径，下载到临时目录
        如果是本地路径，直接返回
        
        Args:
            file_path: 文件路径（MinIO 或本地）
            job_id: 任务ID
            file_type: 文件类型（dwg 或 prt）
        
        Returns:
            本地文件路径
        """
        import tempfile
        from pathlib import Path
        
        # 判断是否为本地路径（以 / 或盘符开头，如 C:）
        if file_path.startswith('/') or (len(file_path) > 1 and file_path[1] == ':'):
            # 本地路径，检查文件是否存在
            if Path(file_path).exists():
                self.logger.info(f"[NCTimeAgent] 使用本地文件: {file_path}")
                return file_path
            else:
                raise FileNotFoundError(f"本地文件不存在: {file_path}")
        
        # MinIO 路径，需要下载
        self.logger.info(f"[NCTimeAgent] 检测到 MinIO 路径，开始下载: {file_path}")
        
        try:
            from scripts.minio_client import MinIOClient
            
            # 创建 MinIO 客户端
            minio_client = MinIOClient()
            
            # 创建临时目录
            temp_dir = Path(tempfile.gettempdir()) / "nc_agent" / job_id
            temp_dir.mkdir(parents=True, exist_ok=True)
            
            # 生成本地文件名
            file_ext = Path(file_path).suffix
            local_file = temp_dir / f"{file_type}{file_ext}"
            
            # 下载文件（使用正确的方法名 get_file）
            success = minio_client.get_file(file_path, str(local_file))
            
            if not success:
                raise Exception(f"从 MinIO 下载文件失败: {file_path}")
            
            self.logger.info(f"[NCTimeAgent] 文件下载成功: {file_path} -> {local_file}")
            return str(local_file)
            
        except Exception as e:
            self.logger.error(f"[NCTimeAgent] 处理文件失败: {file_path}, error={e}")
            raise
    
    async def _cleanup_temp_files(self, temp_files: List[str]):
        """
        清理临时文件
        
        Args:
            temp_files: 临时文件路径列表
        """
        if not temp_files:
            return
        
        import os
        from pathlib import Path
        
        for file_path in temp_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    self.logger.debug(f"[NCTimeAgent] 已删除临时文件: {file_path}")
                    
                    # 尝试删除空的临时目录
                    parent_dir = Path(file_path).parent
                    if parent_dir.exists() and not list(parent_dir.iterdir()):
                        parent_dir.rmdir()
                        self.logger.debug(f"[NCTimeAgent] 已删除空目录: {parent_dir}")
                        
            except Exception as e:
                self.logger.warning(f"[NCTimeAgent] 清理临时文件失败: {file_path}, error={e}")
    
    async def _save_subgraph_response(
        self, 
        job_id: str, 
        subgraph_id: str, 
        subgraph_name: str, 
        subgraph_data: Dict[str, Any]
    ):
        """
        保存单个子图的 NC Agent 响应数据到本地文件
        
        Args:
            job_id: 任务ID
            subgraph_id: 子图完整ID（如 2cd0b581-f2b9-481d-a2fd-33074f57ebd4_PH-01）
            subgraph_name: 子图文件名（如 PH-01-M250297-P5.json）
            subgraph_data: 子图的 NC 数据
        """
        import json
        from pathlib import Path
        from datetime import datetime
        
        try:
            # 创建保存目录：logs/nc_responses/job_id/subgraphs/
            save_dir = Path("logs") / "nc_responses" / job_id / "subgraphs"
            save_dir.mkdir(parents=True, exist_ok=True)
            
            # 提取子图短ID（如 PH-01）
            subgraph_short_id = self._extract_subgraph_id(subgraph_name)
            
            # 构建子图数据结构
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            data = {
                "job_id": job_id,
                "subgraph_id": subgraph_id,
                "subgraph_short_id": subgraph_short_id,
                "file_name": subgraph_name,
                "timestamp": timestamp,
                "meta_data": subgraph_data.get("meta_data", {}),
                "operations": subgraph_data.get("operations", [])
            }
            
            # 保存为单独的 JSON 文件
            filename = f"{subgraph_short_id}_{timestamp}.json"
            filepath = save_dir / filename
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            self.logger.debug(
                f"[NCTimeAgent] 保存子图数据: {subgraph_short_id} -> {filepath}"
            )
            
        except Exception as e:
            self.logger.error(
                f"[NCTimeAgent] 保存子图数据失败: {subgraph_name}, error={e}", 
                exc_info=True
            )
    
    async def _save_raw_response(self, job_id: str, nc_result: Dict[str, Any]):
        """
        保存 NC Agent 原始响应到本地文件（按子图保存）
        
        Args:
            job_id: 任务ID
            nc_result: NC Agent 返回的完整响应
        """
        import json
        from pathlib import Path
        from datetime import datetime
        
        try:
            # 创建保存目录：logs/nc_responses/job_id/
            save_dir = Path("logs") / "nc_responses" / job_id
            save_dir.mkdir(parents=True, exist_ok=True)
            
            # 提取 json_output（包含所有子图数据）
            json_output = nc_result.get("data", {}).get("json_output", {})
            
            if not json_output:
                self.logger.warning(f"[NCTimeAgent] NC Agent 返回空数据，无法保存")
                return
            
            # 构建完整的数据结构（包含所有子图）
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            complete_data = {
                "job_id": job_id,
                "timestamp": timestamp,
                "task_id": nc_result.get("data", {}).get("task_id"),
                "execution_time": nc_result.get("data", {}).get("execution_time"),
                "total_subgraphs": len(json_output),
                "subgraphs": {}
            }
            
            # 遍历每个子图，保存详细数据
            for subgraph_name, subgraph_data in json_output.items():
                # 提取子图短ID（如 PH-01）
                subgraph_short_id = self._extract_subgraph_id(subgraph_name)
                
                # 保存子图数据
                complete_data["subgraphs"][subgraph_short_id] = {
                    "file_name": subgraph_name,
                    "meta_data": subgraph_data.get("meta_data", {}),
                    "operations": subgraph_data.get("operations", [])
                }
            
            # 保存为一个完整的 JSON 文件
            filename = f"nc_data_{job_id}_{timestamp}.json"
            filepath = save_dir / filename
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(complete_data, f, ensure_ascii=False, indent=2)
            
            self.logger.info(
                f"[NCTimeAgent] ✅ NC 数据已保存: {filepath} "
                f"(包含 {len(json_output)} 个子图)"
            )
            
        except Exception as e:
            self.logger.error(f"[NCTimeAgent] 保存 NC 数据失败: {e}", exc_info=True)
    
    def _parse_operations(self, operations: List[Dict[str, Any]], volume_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        旧版 NC JSON 解析逻辑（已停用，保留作参考，不删除）

        解析操作数据，按面编码组织时间信息
        
        新规则（按面编码组织）：
        1. 面编码：使用业务实际编码 ["Z", "B", "C", "C_B", "Z_VIEW", "B_VIEW"]
        2. 钻孔：提取操作名格式为XX_XX_XX时取中间段，XX_ZXZ时取ZXZ
        3. 开粗：累加该面所有含"开粗"的操作Toolpath Time
        4. 精铣：分"半精""全精"两类，分别累加
        5. 空值约束：某面无任何加工类型时，details为空数组
        
        Args:
            operations: 操作列表
            volume_data: 体积数据（可选）
        
        Returns:
            {
                "nc_details": [
                    {"face_code": "Z", "details": [{"code": "开粗", "value": "150.0"}, ...]},
                    {"face_code": "B", "details": [...]},
                    ...
                ],
                "volume_mm3": 12345.678  # 体积（立方毫米）
            }
        """
        # 初始化6个面的数据结构（使用业务实际编码）
        face_codes = ["Z", "B", "C", "C_B", "Z_VIEW", "B_VIEW"]
        face_data = {face: {} for face in face_codes}  # {face_code: {code: total_time}}
        
        # 追踪当前面（只有遇到明确的面标识时才切换）
        current_face = "B_VIEW"  # 默认面
        
        for operation in operations:
            operation_name = operation.get("operation_name", "")
            
            # 提取 Toolpath Time（取 parameters 里面的第一个 value）
            time_value = self._extract_toolpath_time(operation)
            if time_value is None:
                continue
            
            # 判断加工类型
            process_code = None
            
            # 1. 检查操作名是否以面标识开头，如果是则切换当前面
            # 注意：必须先检查复合前缀（Z_VIEW_、B_VIEW_、C_B_），再检查单字母前缀（Z_、B_、X_）
            if "_" in operation_name:
                # 先检查复合前缀
                if operation_name.startswith("Z_VIEW_"):
                    current_face = "Z_VIEW"
                    if self._is_drilling_operation(operation_name):
                        process_code = self._extract_operation_code(operation_name)
                
                elif operation_name.startswith("B_VIEW_"):
                    current_face = "B_VIEW"
                    if self._is_drilling_operation(operation_name):
                        process_code = self._extract_operation_code(operation_name)
                
                elif operation_name.startswith("C_B_"):
                    current_face = "C_B"
                    if self._is_drilling_operation(operation_name):
                        process_code = self._extract_operation_code(operation_name)
                
                # 再检查单字母前缀
                else:
                    prefix = operation_name.split("_")[0]
                    
                    if prefix == "Z":
                        current_face = "Z"
                        if self._is_drilling_operation(operation_name):
                            process_code = self._extract_operation_code(operation_name)
                    
                    elif prefix == "B":
                        current_face = "B"
                        if self._is_drilling_operation(operation_name):
                            process_code = self._extract_operation_code(operation_name)
                    
                    elif prefix == "X":
                        current_face = "C"
                        if self._is_drilling_operation(operation_name):
                            process_code = self._extract_operation_code(operation_name)
            
            # 2. 检查是否包含"粗"字（开粗）- 归属到当前面
            if not process_code and "粗" in operation_name:
                process_code = "开粗"
            
            # 3. 检查是否包含"精"字（半精、全精）- 归属到当前面
            elif not process_code and "精" in operation_name:
                if "半精" in operation_name:
                    process_code = "半精"
                elif "全精" in operation_name:
                    process_code = "全精"
                else:
                    process_code = "精铣"
            
            # 如果没有匹配到任何类型，跳过
            if not process_code:
                self.logger.warning(f"[NCTimeAgent] 无法分类操作: {operation_name}")
                continue
            
            # 汇总到当前面和对应的代码
            if process_code not in face_data[current_face]:
                face_data[current_face][process_code] = Decimal("0")
            face_data[current_face][process_code] += time_value
        
        # 构建 nc_details（按面编码组织）
        nc_details = []
        
        for face_code in face_codes:
            details = []
            
            # 获取该面的所有加工类型
            face_processes = face_data[face_code]
            
            # 只添加有值的加工类型（不添加0值项）
            for code, value in face_processes.items():
                if value > 0:
                    details.append({
                        "code": code,
                        "value": value  # 保留原始浮点精度
                    })
            
            # 添加面数据（即使details为空也要添加）
            nc_details.append({
                "face_code": face_code,
                "details": details
            })
        
        result = {
            "nc_details": nc_details
        }
        
        # 添加体积数据（如果有）
        if volume_data:
            volume_mm3 = volume_data.get("volume_mm3")
            if volume_mm3 is not None:
                result["volume_mm3"] = volume_mm3
        
        return result
    
    def _extract_toolpath_time(self, operation: Dict[str, Any]) -> Decimal:
        """
        从操作中提取 Toolpath Time（取 parameters 里面的第一个 value）
        
        Args:
            operation: 操作数据
        
        Returns:
            时间值（分钟），如果未找到返回 None
        """
        parameters = operation.get("parameters", [])
        
        # 取第一个参数的 value（通常是 id=124 的 Toolpath Time）
        if parameters and len(parameters) > 0:
            first_param = parameters[0]
            value = first_param.get("value")
            if value is not None:
                return Decimal(str(value))
        
        return None
    
    def _is_drilling_operation(self, operation_name: str) -> bool:
        """
        判断是否为钻孔操作
        
        规则：
        1. XX_ZXZ 格式（如 Z_ZXZ）
        2. XX_XX_XX 格式（如 Z_M_A14, Z_L_A3, B_M1_A9, X_C_D17.0）
        
        Args:
            operation_name: 操作名称
        
        Returns:
            是否为钻孔操作
        """
        # 检查是否为 XX_ZXZ 格式
        if re.match(r"^[A-Z]+_ZXZ$", operation_name):
            return True
        
        # 检查是否为 XX_XX_XX 格式（中间部分是字母+可选数字的组合）
        # 例如：Z_M_A14, Z_L_A3, B_M1_A9, Z_C_A18, X_C_D17.0
        if re.match(r"^[A-Z]+_[A-Z]\d*_", operation_name):
            return True
        
        return False
    
    def _extract_operation_code(self, operation_name: str) -> str:
        """
        从操作名称中提取代码
        
        规则：
        1. 对于 XX_XX_XX 格式，提取中间的 XX（可能是字母+数字，如 M, L, M1, C）
        2. 对于 XX_ZXZ 格式，提取 ZXZ
        
        例如：
        - Z_M_A14 -> M
        - Z_L_A3 -> L
        - B_M1_A9 -> M1
        - Z_ZXZ -> ZXZ
        - Z_C_A18 -> C
        - X_C_D17.0 -> C
        
        Args:
            operation_name: 操作名称
        
        Returns:
            操作代码（如 M, L, M1, C, ZXZ）
        """
        # 检查是否为 XX_ZXZ 格式
        if re.match(r"^[A-Z]+_ZXZ$", operation_name):
            return "ZXZ"
        
        # 提取 XX_XX_XX 格式中间的 XX（字母+可选数字）
        # 例如：Z_M_A14 -> M, Z_M1_A18 -> M1, Z_C_A18 -> C, X_C_D17.0 -> C
        match = re.search(r"^[A-Z]+_([A-Z]\d*)_", operation_name)
        if match:
            return match.group(1)
        
        # 如果无法匹配，返回原始名称
        self.logger.warning(f"[NCTimeAgent] 无法提取操作代码: {operation_name}")
        return operation_name
    
    async def _save_nc_time_data(self, subgraph_id: str, time_data: Dict[str, Any]):
        """
        保存 NC 时间数据到数据库
        
        Args:
            subgraph_id: 子图ID
            time_data: 时间数据，格式：
                {
                    "nc_details": [
                        {"face_code": "A", "details": [{"code": "开粗", "value": 150.0}, ...]},
                        ...
                    ],
                    "volume_mm3": 12345.678  # 可选
                }
        """
        async for db in get_db():
            # 转换 Decimal 类型为 float（JSON 序列化需要）
            def convert_decimals(obj):
                """递归转换 Decimal 为 float"""
                from decimal import Decimal
                if isinstance(obj, Decimal):
                    return float(obj)
                elif isinstance(obj, dict):
                    return {k: convert_decimals(v) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [convert_decimals(item) for item in obj]
                return obj

            actual_times = self._summarize_actual_nc_times(time_data.get("nc_details", []))
            
            # 准备更新数据
            nc_time_cost_data = convert_decimals({"nc_details": time_data["nc_details"]})
            update_values = {
                "nc_time_cost": nc_time_cost_data,
                "created_at": datetime.utcnow()
            }
            
            # 如果有体积数据，也一起更新
            if "volume_mm3" in time_data:
                update_values["volume_mm3"] = (
                    float(time_data["volume_mm3"])
                    if isinstance(time_data["volume_mm3"], Decimal)
                    else time_data["volume_mm3"]
                )
            
            # 更新 features 表（nc_time_cost 和 volume_mm3 字段）
            await db.execute(
                update(Feature)
                .where(Feature.subgraph_id == subgraph_id)
                .values(**update_values)
            )

            await db.execute(
                update(Subgraph)
                .where(Subgraph.subgraph_id == subgraph_id)
                .values(
                    nc_roughing_time=actual_times["nc_roughing_time"],
                    nc_milling_time=actual_times["nc_milling_time"],
                    drilling_time=actual_times["drilling_time"],
                )
            )
            
            await db.commit()
            
            self.logger.debug(
                f"[NCTimeAgent] 保存 NC 时间数据到 features 表: subgraph_id={subgraph_id}, "
                f"faces={len(time_data['nc_details'])}, "
                f"volume_mm3={time_data.get('volume_mm3', 'N/A')}, "
                f"roughing_h={actual_times['nc_roughing_time']}, "
                f"milling_h={actual_times['nc_milling_time']}, "
                f"drilling_h={actual_times['drilling_time']}"
            )
            
            break

    async def _save_subgraph_processing_description(
        self,
        subgraph_id: str,
        processing_text: Any,
    ):
        """将 NC JSON 中的原始 processing 文本暂存到 subgraphs.process_description。"""
        normalized_text = str(processing_text or "").strip()
        if not normalized_text:
            return

        async for db in get_db():
            await db.execute(
                update(Subgraph)
                .where(Subgraph.subgraph_id == subgraph_id)
                .values(process_description=normalized_text)
            )
            await db.commit()

            self.logger.debug(
                f"[NCTimeAgent] 保存原始 processing 到 subgraphs.process_description: "
                f"subgraph_id={subgraph_id}, processing={normalized_text}"
            )
            break

    def _summarize_actual_nc_times(self, nc_details: List[Dict[str, Any]]) -> Dict[str, float]:
        """汇总 NC 原始明细中的开粗/精铣/钻床实际加工时间，不乘数量。"""
        roughing_minutes = Decimal("0")
        milling_minutes = Decimal("0")
        drilling_minutes = Decimal("0")

        for face_detail in nc_details or []:
            details = face_detail.get("details", []) or []
            for detail in details:
                code = str(detail.get("code") or "").strip()
                value = detail.get("value")
                try:
                    minutes = Decimal(str(value))
                except Exception:
                    continue

                if code in ["精铣", "半精", "全精"]:
                    milling_minutes += minutes
                elif code == "开粗":
                    roughing_minutes += minutes
                else:
                    drilling_minutes += minutes

        return {
            "nc_roughing_time": round(float(roughing_minutes / Decimal("60")), 2),
            "nc_milling_time": round(float(milling_minutes / Decimal("60")), 2),
            "drilling_time": round(float(drilling_minutes / Decimal("60")), 2),
        }

    async def _save_nc_volume_data(self, subgraph_id: str, volume_mm3: Any):
        if volume_mm3 in (None, ""):
            return

        async for db in get_db():
            await db.execute(
                update(Feature)
                .where(Feature.subgraph_id == subgraph_id)
                .values(
                    volume_mm3=float(volume_mm3),
                    created_at=datetime.utcnow()
                )
            )
            await db.commit()
            self.logger.debug(
                f"[NCTimeAgent] 保存 NC 体积数据到 features 表: subgraph_id={subgraph_id}, volume_mm3={volume_mm3}"
            )
            break

    async def _process_nc_debug_metadata(
        self,
        job_id: str,
        nc_result: Dict[str, Any],
        source: str = "minio_json_debug",
    ):
        fail_itemcodes = self._extract_fail_itemcodes(nc_result)
        await self._save_nc_failed_itemcodes(job_id, fail_itemcodes)

        json_output = nc_result.get("data", {}).get("json_output", {}) or {}
        volume_saved_count = 0

        for subgraph_name, subgraph_data in json_output.items():
            try:
                subgraph_short_id = self._extract_subgraph_id(subgraph_name)
                subgraph_id = await self._find_subgraph_id(job_id, subgraph_short_id)
                if not subgraph_id:
                    continue

                volume_mm3 = (
                    (subgraph_data.get("batch_meta", {}) or {})
                    .get("volume_data", {})
                    .get("volume_mm3")
                )
                if volume_mm3 in (None, ""):
                    continue

                await self._save_nc_volume_data(subgraph_id, volume_mm3)
                volume_saved_count += 1
            except Exception as exc:
                self.logger.error(
                    f"[NCTimeAgent] 处理 debug JSON 体积失败: source={source}, subgraph={subgraph_name}, error={exc}",
                    exc_info=True,
                )

        self.logger.info(
            f"[NCTimeAgent] debug JSON 元数据处理完成: source={source}, "
            f"fail_itemcodes={len(fail_itemcodes)}, volume_saved={volume_saved_count}"
        )

    async def _process_nc_excel_workbooks(
        self,
        job_id: str,
        context: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """旧版 Excel NC 解析入口（已停用，保留作参考，不删除）"""
        mold_code = await self._get_job_mold_code(job_id)
        subgraphs = await self._get_job_subgraphs(job_id)
        self._log_visible(
            "info",
            f"[NCTimeAgent] NC Excel 分支开始: job_id={job_id}, mold_code={mold_code or '<empty>'}, "
            f"subgraph_count={len(subgraphs)}"
        )

        if not subgraphs:
            self.logger.warning(f"[NCTimeAgent] job 没有子图记录，跳过 Excel NC 解析: job_id={job_id}")
            return {
                "status": "error",
                "message": "NC 时间计算失败: 当前任务没有子图数据"
            }

        search_dirs = self._get_nc_excel_search_dirs(context)
        downloaded_excel_dir = self._download_nc_excel_files_to_logs(job_id, mold_code)
        if downloaded_excel_dir and downloaded_excel_dir not in search_dirs:
            search_dirs.insert(0, downloaded_excel_dir)
        explicit_excel_path = self._normalize_excel_file_path(context.get("nc_excel_path"))
        self._log_visible(
            "info",
            f"[NCTimeAgent] NC Excel 搜索路径: downloaded_excel_dir={downloaded_excel_dir}, "
            f"explicit_excel_path={explicit_excel_path}, search_dirs={[str(item) for item in search_dirs]}"
        )

        success_count = 0
        failed_count = 0

        for subgraph in subgraphs:
            subgraph_id = subgraph["subgraph_id"]
            part_code = (subgraph.get("part_code") or "").strip()
            if not part_code:
                self.logger.warning(f"[NCTimeAgent] 子图缺少 part_code，跳过 Excel NC 解析: subgraph_id={subgraph_id}")
                failed_count += 1
                continue

            workbook_path = self._find_nc_excel_workbook(
                part_code=part_code,
                mold_code=mold_code,
                search_dirs=search_dirs,
                explicit_excel_path=explicit_excel_path,
            )
            if not workbook_path:
                self.logger.warning(
                    f"[NCTimeAgent] 未找到 NC Excel 文件: part_code={part_code}, mold_code={mold_code}, "
                    f"search_dirs={[str(item) for item in search_dirs]}"
                )
                failed_count += 1
                continue

            try:
                self.logger.info(
                    f"[NCTimeAgent] 使用 Excel 解析 NC: subgraph_id={subgraph_id}, part_code={part_code}, "
                    f"workbook={workbook_path}"
                )
                time_data = self._parse_nc_excel_workbook(workbook_path)
                await self._save_nc_time_data(subgraph_id, time_data)
                success_count += 1
            except Exception as exc:
                self.logger.error(
                    f"[NCTimeAgent] Excel NC 解析失败: subgraph_id={subgraph_id}, part_code={part_code}, "
                    f"workbook={workbook_path}, error={exc}",
                    exc_info=True
                )
                failed_count += 1

        if success_count == 0:
            self.logger.info(f"[NCTimeAgent] 未命中任何可用 Excel NC 文件，继续走旧 NC JSON 流程: job_id={job_id}")
            return None

        total_count = success_count + failed_count
        if self.progress_publisher:
            from shared.progress_stages import ProgressStage, ProgressPercent
            self.progress_publisher.publish_progress(
                job_id=job_id,
                stage=ProgressStage.NC_CALCULATION_COMPLETED,
                progress=ProgressPercent.NC_CALCULATION_COMPLETED,
                message=f"NC Excel 解析完成，成功 {success_count}/{total_count}",
                details={
                    "source": "excel",
                    "total_subgraphs": total_count,
                    "success_count": success_count,
                    "failed_count": failed_count
                }
            )

        return {
            "status": "ok",
            "message": f"NC Excel 解析完成，成功 {success_count}/{total_count}",
            "summary": {
                "source": "excel",
                "total_subgraphs": total_count,
                "success_count": success_count,
                "failed_count": failed_count
            }
        }

    async def _process_nc_json_from_minio(self, job_id: str) -> Optional[Dict[str, Any]]:
        mold_code = await self._get_job_mold_code(job_id)
        self._log_visible(
            "info",
            f"[NCTimeAgent] NC MinIO JSON 分支开始: job_id={job_id}, mold_code={mold_code or '<empty>'}"
        )
        if not mold_code:
            self.logger.warning(f"[NCTimeAgent] 缺少模号，跳过 MinIO JSON 解析: job_id={job_id}")
            return None

        minio_prefix = self._query_nc_excel_minio_prefix(mold_code)
        self._log_visible(
            "info",
            f"[NCTimeAgent] NC MinIO JSON 查询路径: mold_code={mold_code}, minio_prefix={minio_prefix or '<empty>'}"
        )
        if not minio_prefix:
            self.logger.warning(f"[NCTimeAgent] 外部 NC 库未查到 MinIO 路径，跳过 JSON 解析: mold_code={mold_code}")
            return None

        try:
            objects = self._list_nc_minio_objects(minio_prefix)
            self._log_visible(
                "info",
                f"[NCTimeAgent] NC MinIO 对象枚举完成: mold_code={mold_code}, path={minio_prefix}, "
                f"object_count={len(objects)}"
            )
            if objects:
                self._log_visible(
                    "info",
                    f"[NCTimeAgent] NC MinIO 对象列表: {objects}"
                )
        except Exception as exc:
            self.logger.error(
                f"[NCTimeAgent] 枚举 NC MinIO 对象失败: mold_code={mold_code}, path={minio_prefix}, error={exc}",
                exc_info=True,
            )
            return None

        debug_json_objects = sorted(
            obj for obj in objects
            if "/nc_agent_debug/" in obj.replace("\\", "/") and obj.lower().endswith(".json")
        )
        self._log_visible(
            "info",
            f"[NCTimeAgent] NC MinIO debug JSON 命中数: {len(debug_json_objects)}"
        )
        local_debug_files: List[Path] = []
        if debug_json_objects:
            debug_download_root = self.logs_root / "nc_agent_debug"
            debug_download_root.mkdir(parents=True, exist_ok=True)
            self._log_visible(
                "info",
                f"[NCTimeAgent] NC MinIO debug JSON 下载列表: {debug_json_objects}"
            )
            for object_name in debug_json_objects:
                local_path = debug_download_root / Path(object_name).name
                self._download_nc_minio_object(object_name, local_path)
                local_debug_files.append(local_path)

        subgraph_json_objects = sorted(
            obj for obj in objects
            if "/nc_responses/" in obj.replace("\\", "/")
            and "/subgraphs/" in obj.replace("\\", "/")
            and obj.lower().endswith(".json")
        )
        self._log_visible(
            "info",
            f"[NCTimeAgent] NC MinIO subgraph JSON 命中数: {len(subgraph_json_objects)}"
        )
        local_subgraph_files: List[Path] = []
        if subgraph_json_objects:
            download_root = self.logs_root / "nc_responses" / job_id / "subgraphs"
            download_root.mkdir(parents=True, exist_ok=True)
            self._log_visible(
                "info",
                f"[NCTimeAgent] NC MinIO subgraph JSON 下载列表: {subgraph_json_objects}"
            )
            for object_name in subgraph_json_objects:
                local_path = download_root / Path(object_name).name
                self._download_nc_minio_object(object_name, local_path)
                local_subgraph_files.append(local_path)

        debug_nc_result: Optional[Dict[str, Any]] = None
        if local_debug_files:
            latest_debug_file = max(local_debug_files, key=lambda item: item.stat().st_mtime)
            self._log_visible(
                "info",
                f"[NCTimeAgent] 使用 MinIO debug JSON 解析 NC 元数据: {latest_debug_file}"
            )
            with open(latest_debug_file, "r", encoding="utf-8") as f:
                debug_nc_result = json.load(f)
            await self._process_nc_debug_metadata(job_id, debug_nc_result, source="minio_json_debug")

        if local_subgraph_files:
            self._log_visible(
                "info",
                f"[NCTimeAgent] 使用 MinIO subgraph JSON 解析 NC: job_id={job_id}, files={len(local_subgraph_files)}"
            )
            return await self._process_nc_subgraph_json_files(
                job_id, local_subgraph_files, source="minio_json_subgraphs"
            )

        if debug_nc_result is not None:
            # 旧逻辑保留作参考：此前 MinIO 仅命中 debug JSON 时，会回退到
            # _process_nc_result_payload() 继续走旧版全量解析。
            # self.logger.warning(
            #     "[NCTimeAgent] MinIO 仅命中 debug JSON，未命中 subgraph JSON，回退到 debug JSON 全量解析"
            # )
            # return await self._process_nc_result_payload(job_id, debug_nc_result, source="minio_json_debug")
            self.logger.error(
                "[NCTimeAgent] MinIO 仅命中 debug JSON，未命中 subgraph JSON；旧版 debug 全量解析已停用"
            )
            return {
                "status": "error",
                "message": "NC 时间计算失败: MinIO 仅命中 debug JSON，缺少 subgraph JSON"
            }

        self.logger.error(
            f"[NCTimeAgent] MinIO 路径下未命中可用的 subgraph JSON: mold_code={mold_code}, path={minio_prefix}"
        )
        return None

    async def _process_nc_result_payload(
        self,
        job_id: str,
        nc_result: Dict[str, Any],
        source: str = "http",
    ) -> Dict[str, Any]:
        code = nc_result.get("code")
        message = nc_result.get("message", "")

        if code == 500:
            error_msg = nc_result.get("message", "NC Agent 调用失败")
            self.logger.error(f"[NCTimeAgent] NC 数据源完全失败({source}): {error_msg}")

            if self.progress_publisher:
                from shared.progress_stages import ProgressStage, ProgressPercent
                self.progress_publisher.publish_progress(
                    job_id=job_id,
                    stage=ProgressStage.NC_CALCULATION_FAILED,
                    progress=ProgressPercent.NC_CALCULATION_STARTED,
                    message=f"NC 时间计算失败: {error_msg}",
                    details={"error": error_msg, "code": 500, "source": source}
                )

            return {"status": "error", "message": error_msg}
        elif code == 206:
            self.logger.warning(f"[NCTimeAgent] NC 数据源部分成功({source}): {message}")
        elif code == 200:
            self.logger.info(f"[NCTimeAgent] NC 数据源成功({source}): {message}")
        else:
            self.logger.warning(f"[NCTimeAgent] NC 数据源返回未知代码({source}) {code}: {message}")

        json_output = nc_result.get("data", {}).get("json_output", {})
        if not json_output:
            error_details = nc_result.get("data", {}).get("error_details", "未知错误")
            failed_step = nc_result.get("data", {}).get("failed_at_step", "未知步骤")

            self.logger.warning(f"[NCTimeAgent] NC 数据源返回空数据({source})")
            self.logger.warning(f"[NCTimeAgent] 失败步骤: Step {failed_step}")
            self.logger.warning(f"[NCTimeAgent] 错误详情: {error_details[:500]}")

            if self.progress_publisher:
                from shared.progress_stages import ProgressStage, ProgressPercent
                self.progress_publisher.publish_progress(
                    job_id=job_id,
                    stage=ProgressStage.NC_CALCULATION_FAILED,
                    progress=ProgressPercent.NC_CALCULATION_STARTED,
                    message=f"NC 时间计算失败: NC 数据源返回空数据 (Step {failed_step})",
                    details={
                        "error": f"NC 数据源返回空数据 (Step {failed_step})",
                        "failed_step": failed_step,
                        "error_details": error_details[:500] if error_details else "未知错误",
                        "source": source,
                    }
                )

            return {
                "status": "error",
                "message": f"NC 数据处理失败 (Step {failed_step})",
                "details": {
                    "failed_step": failed_step,
                    "error": error_details[:500] if error_details else "未知错误",
                    "source": source,
                }
            }

        self.logger.info(f"[NCTimeAgent] ========================================")
        self.logger.info(f"[NCTimeAgent] NC 返回数据（原始）source={source}")
        self.logger.info(f"[NCTimeAgent] 子图总数: {len(json_output)}")
        self.logger.info(f"[NCTimeAgent] ========================================")

        for subgraph_name, subgraph_data in json_output.items():
            self.logger.info(f"[NCTimeAgent] ")
            self.logger.info(f"[NCTimeAgent] 子图: {subgraph_name}")
            operations = subgraph_data.get("operations", [])
            self.logger.info(f"[NCTimeAgent] 操作数量: {len(operations)}")

            for op in operations:
                op_name = op.get("operation_name", "")
                params = op.get("parameters", [])
                time_param = next((p for p in params if p.get("id") == 124), None)

                if time_param:
                    time_value = time_param.get("value", 0)
                    display_name = time_param.get("display_name", "")
                    self.logger.info(
                        f"[NCTimeAgent]   操作: {op_name} | 时间: {time_value} | "
                        f"参数: id={time_param.get('id')}, {display_name}"
                    )
                else:
                    self.logger.warning(f"[NCTimeAgent]   操作: {op_name} | 未找到 id=124 的时间参数")

        self.logger.info(f"[NCTimeAgent] ========================================")

        success_count = 0
        failed_count = 0

        for subgraph_name, subgraph_data in json_output.items():
            try:
                subgraph_short_id = self._extract_subgraph_id(subgraph_name)
                subgraph_id = await self._find_subgraph_id(job_id, subgraph_short_id)

                if not subgraph_id:
                    self.logger.warning(
                        f"[NCTimeAgent] 未找到子图映射: {subgraph_name} -> {subgraph_short_id}"
                    )
                    failed_count += 1
                    continue

                if source == "http":
                    await self._save_subgraph_response(job_id, subgraph_id, subgraph_name, subgraph_data)
                else:
                    self.logger.info(
                        f"[NCTimeAgent] 跳过本地子图 JSON 落盘: source={source}, "
                        f"subgraph={subgraph_name}, subgraph_id={subgraph_id}"
                    )
                operations = subgraph_data.get("operations", [])
                # 旧逻辑保留作参考：此前完整响应分支仍会走 _parse_operations()，
                # 其面别判断依赖 Z/B/X/Z_VIEW/B_VIEW/C_B 等旧前缀规则。
                # volume_data = subgraph_data.get("batch_meta", {}).get("volume_data", {})
                # time_data = self._parse_operations(operations, volume_data)

                # 新逻辑：统一按 parent_group -> 面别映射解析，避免完整响应分支继续使用旧前缀切面。
                time_data = self._parse_subgraph_json_operations(operations)
                await self._save_subgraph_processing_description(
                    subgraph_id,
                    (subgraph_data.get("meta_data") or {}).get("processing"),
                )
                await self._save_nc_time_data(subgraph_id, time_data)

                success_count += 1
                self.logger.info(f"[NCTimeAgent] 成功处理子图: {subgraph_name} -> {subgraph_id}")
            except Exception as e:
                self.logger.error(f"[NCTimeAgent] 处理子图失败: {subgraph_name}, error={e}", exc_info=True)
                failed_count += 1

        total_count = success_count + failed_count
        self.logger.info(
            f"[NCTimeAgent] NC 时间计算完成: total={total_count}, success={success_count}, "
            f"failed={failed_count}, source={source}"
        )

        if self.progress_publisher:
            from shared.progress_stages import ProgressStage, ProgressPercent
            self.progress_publisher.publish_progress(
                job_id=job_id,
                stage=ProgressStage.NC_CALCULATION_COMPLETED,
                progress=ProgressPercent.NC_CALCULATION_COMPLETED,
                message=f"NC 时间计算完成，成功 {success_count}/{total_count}",
                details={
                    "total_subgraphs": total_count,
                    "success_count": success_count,
                    "failed_count": failed_count,
                    "source": source,
                }
            )

        return {
            "status": "ok",
            "message": f"NC 时间计算完成，成功 {success_count}/{total_count}",
            "summary": {
                "total_subgraphs": total_count,
                "success_count": success_count,
                "failed_count": failed_count,
                "source": source,
            }
        }

    async def _process_nc_subgraph_json_files(
        self,
        job_id: str,
        json_files: List[Path],
        source: str = "minio_json_subgraphs",
    ) -> Dict[str, Any]:
        success_count = 0
        failed_count = 0

        for json_file in json_files:
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                subgraph_short_id = (
                    data.get("subgraph_short_id")
                    or self._extract_subgraph_id(data.get("file_name", json_file.name))
                )
                subgraph_id = data.get("subgraph_id") or await self._find_subgraph_id(job_id, subgraph_short_id)
                if not subgraph_id:
                    self.logger.warning(
                        f"[NCTimeAgent] 未找到子图映射(JSON): file={json_file}, short_id={subgraph_short_id}"
                    )
                    failed_count += 1
                    continue

                operations = data.get("operations", [])
                time_data = self._parse_subgraph_json_operations(operations)
                await self._save_subgraph_processing_description(
                    subgraph_id,
                    (data.get("meta_data") or {}).get("processing"),
                )
                await self._save_nc_time_data(subgraph_id, time_data)

                success_count += 1
                self.logger.info(f"[NCTimeAgent] 成功处理子图(JSON): {json_file.name} -> {subgraph_id}")
            except Exception as exc:
                self.logger.error(
                    f"[NCTimeAgent] 处理 MinIO JSON 子图失败: file={json_file}, error={exc}",
                    exc_info=True,
                )
                failed_count += 1

        total_count = success_count + failed_count
        if self.progress_publisher:
            from shared.progress_stages import ProgressStage, ProgressPercent
            self.progress_publisher.publish_progress(
                job_id=job_id,
                stage=ProgressStage.NC_CALCULATION_COMPLETED,
                progress=ProgressPercent.NC_CALCULATION_COMPLETED,
                message=f"NC JSON 解析完成，成功 {success_count}/{total_count}",
                details={
                    "total_subgraphs": total_count,
                    "success_count": success_count,
                    "failed_count": failed_count,
                    "source": source,
                }
            )

        return {
            "status": "ok",
            "message": f"NC JSON 解析完成，成功 {success_count}/{total_count}",
            "summary": {
                "total_subgraphs": total_count,
                "success_count": success_count,
                "failed_count": failed_count,
                "source": source,
            }
        }

    def _parse_subgraph_json_operations(self, operations: List[Dict[str, Any]]) -> Dict[str, Any]:
        nc_details = []
        face_details_map: Dict[str, List[Dict[str, Any]]] = {
            face_code: [] for face_code in self.JSON_FACE_CODE_MAP.values()
        }

        for operation in operations:
            operation_name = (operation.get("operation_name") or "").strip()
            if not operation_name:
                continue

            face_code = self._extract_json_face_code(operation)
            if not face_code:
                self.logger.warning(f"[NCTimeAgent] JSON 操作无法识别面别: {operation_name}")
                continue

            time_value = self._extract_parameter_decimal(operation, 124)
            if time_value is None:
                self.logger.warning(f"[NCTimeAgent] JSON 操作缺少 Toolpath Time(id=124): {operation_name}")
                continue

            code = self._extract_excel_operation_code(operation_name)
            if not code:
                self.logger.warning(f"[NCTimeAgent] JSON 操作无法提取工序 code: {operation_name}")
                continue

            face_details_map[face_code].append({
                "code": code,
                "value": time_value,
                "program_name": operation_name,
                "tool": self._extract_json_tool_name(operation_name),
            })

        for face_code in self.JSON_FACE_CODE_MAP.values():
            nc_details.append({
                "face_code": face_code,
                "details": face_details_map.get(face_code, [])
            })

        return {"nc_details": nc_details}

    async def _get_job_mold_code(self, job_id: str) -> str:
        import uuid

        async for db in get_db():
            job = await db.get(Job, uuid.UUID(job_id))
            if not job:
                break

            for filename in [job.dwg_file_name, job.prt_file_name]:
                mold_code = self._extract_mold_code_from_filename(filename)
                if mold_code:
                    return mold_code
            break
        return ""

    async def _get_job_subgraphs(self, job_id: str) -> List[Dict[str, str]]:
        import uuid

        async for db in get_db():
            result = await db.execute(
                select(Subgraph.subgraph_id, Subgraph.part_code)
                .where(Subgraph.job_id == uuid.UUID(job_id))
                .order_by(Subgraph.part_code, Subgraph.subgraph_id)
            )
            return [
                {"subgraph_id": row[0], "part_code": row[1]}
                for row in result.fetchall()
            ]
        return []

    def _extract_mold_code_from_filename(self, filename: Optional[str]) -> str:
        if not filename:
            return ""

        text = Path(str(filename)).name
        match = re.search(r"(M\d+-P\d+)", text, re.IGNORECASE)
        if match:
            return match.group(1).upper()

        # Fallback to the filename stem for cases like "附图2.dwg" or "3d.dwg".
        return Path(text).stem.strip()

    def _get_nc_excel_search_dirs(self, context: Dict[str, Any]) -> List[Path]:
        candidates: List[Path] = []

        for raw_value in [
            context.get("nc_excel_dir"),
            self.nc_excel_dir,
            str(self.nc_excel_logs_dir),
            str(Path.cwd()),
        ]:
            if not raw_value:
                continue
            path = Path(str(raw_value)).expanduser()
            if path.exists() and path.is_dir() and path not in candidates:
                candidates.append(path)

        return candidates

    def _download_nc_excel_files_to_logs(self, job_id: str, mold_code: str) -> Optional[Path]:
        if not mold_code:
            self.logger.warning(f"[NCTimeAgent] 缺少模号，无法从 NC MinIO 下载 Excel: job_id={job_id}")
            return None

        minio_prefix = self._query_nc_excel_minio_prefix(mold_code)
        self._log_visible(
            "info",
            f"[NCTimeAgent] NC Excel 查询路径: job_id={job_id}, mold_code={mold_code}, "
            f"minio_prefix={minio_prefix or '<empty>'}"
        )
        if not minio_prefix:
            self.logger.warning(f"[NCTimeAgent] 外部 NC 库未查到 Excel 路径: mold_code={mold_code}")
            return None

        download_root = self.nc_excel_logs_dir / job_id
        download_root.mkdir(parents=True, exist_ok=True)

        try:
            objects = self._list_nc_minio_objects(minio_prefix)
            xlsx_objects = [obj for obj in objects if obj.lower().endswith(".xlsx")]
            self._log_visible(
                "info",
                f"[NCTimeAgent] NC Excel 枚举结果: path={minio_prefix}, object_count={len(objects)}, "
                f"xlsx_count={len(xlsx_objects)}"
            )
            if xlsx_objects:
                self._log_visible(
                    "info",
                    f"[NCTimeAgent] NC Excel 下载列表: {xlsx_objects}"
                )
            if not xlsx_objects:
                self.logger.warning(
                    f"[NCTimeAgent] NC MinIO 路径下未找到 .xlsx 文件: bucket={os.getenv('NC_SOURCE_MINIO_BUCKET', 'ncresult')}, "
                    f"path={minio_prefix}"
                )
                return None

            self.logger.info(
                f"[NCTimeAgent] NC Excel 下载开始: bucket={os.getenv('NC_SOURCE_MINIO_BUCKET', 'ncresult')}, "
                f"path={minio_prefix}, files={len(xlsx_objects)}"
            )

            for object_name in xlsx_objects:
                relative_name = object_name[len(minio_prefix):].lstrip("/\\") if object_name.startswith(minio_prefix) else Path(object_name).name
                local_path = download_root / Path(relative_name)
                local_path.parent.mkdir(parents=True, exist_ok=True)
                self._download_nc_minio_object(object_name, local_path)

            return download_root
        except Exception as exc:
            self.logger.error(
                f"[NCTimeAgent] 下载 NC Excel 失败: mold_code={mold_code}, path={minio_prefix}, error={exc}",
                exc_info=True
            )
            return None

    def _query_nc_excel_minio_prefix(self, mold_code: str) -> Optional[str]:
        import psycopg2

        query = f"SELECT url FROM {os.getenv('NC_SOURCE_TABLE', 'nc')} WHERE code = %s ORDER BY url LIMIT 1"
        conn = None
        try:
            conn = psycopg2.connect(
                host=os.getenv("NC_SOURCE_DB_HOST"),
                port=int(os.getenv("NC_SOURCE_DB_PORT", "5432")),
                dbname=os.getenv("NC_SOURCE_DB_NAME"),
                user=os.getenv("NC_SOURCE_DB_USER"),
                password=os.getenv("NC_SOURCE_DB_PASSWORD"),
            )
            with conn.cursor() as cursor:
                cursor.execute(query, (mold_code,))
                row = cursor.fetchone()
                if not row or not row[0]:
                    return None
                return str(row[0]).strip().replace("\\", "/").strip("/")
        finally:
            if conn:
                conn.close()

    def _create_nc_minio_client(self):
        from minio import Minio

        endpoint = os.getenv("NC_SOURCE_MINIO_ENDPOINT")
        access_key = os.getenv("NC_SOURCE_MINIO_ACCESS_KEY")
        secret_key = os.getenv("NC_SOURCE_MINIO_SECRET_KEY")
        region = os.getenv("NC_SOURCE_MINIO_REGION", "us-east-1")
        secure = os.getenv("NC_SOURCE_MINIO_USE_HTTPS", "false").lower() == "true"

        return Minio(
            endpoint,
            access_key=access_key,
            secret_key=secret_key,
            secure=secure,
            region=region,
        )

    def _list_nc_minio_objects(self, prefix: str) -> List[str]:
        client = self._create_nc_minio_client()
        bucket = os.getenv("NC_SOURCE_MINIO_BUCKET", "ncresult")
        normalized_prefix = prefix.strip().replace("\\", "/").strip("/")
        if normalized_prefix and not normalized_prefix.endswith("/"):
            normalized_prefix = normalized_prefix + "/"

        return [
            obj.object_name
            for obj in client.list_objects(bucket, prefix=normalized_prefix, recursive=True)
            if getattr(obj, "object_name", None)
        ]

    def _download_nc_minio_object(self, object_name: str, local_path: Path):
        client = self._create_nc_minio_client()
        bucket = os.getenv("NC_SOURCE_MINIO_BUCKET", "ncresult")
        self._log_visible("info", f"[NCTimeAgent] MinIO 下载文件: bucket={bucket}, path={object_name}, local_path={local_path}")
        client.fget_object(
            bucket_name=bucket,
            object_name=object_name,
            file_path=str(local_path)
        )

    def _normalize_excel_file_path(self, excel_path: Optional[str]) -> Optional[Path]:
        if not excel_path:
            return None

        path = Path(str(excel_path)).expanduser()
        if path.exists() and path.is_file():
            return path
        return None

    def _find_nc_excel_workbook(
        self,
        part_code: str,
        mold_code: str,
        search_dirs: List[Path],
        explicit_excel_path: Optional[Path] = None
    ) -> Optional[Path]:
        normalized_part_code = part_code.strip().upper()

        candidate_names = []
        if mold_code:
            candidate_names.append(f"{part_code}-{mold_code}.xlsx")
        candidate_names.append(f"{part_code}.xlsx")

        for search_dir in search_dirs:
            for candidate_name in candidate_names:
                candidate_path = search_dir / candidate_name
                if candidate_path.exists():
                    return candidate_path

            glob_patterns = [f"{part_code}-*.xlsx", f"{part_code}_*.xlsx", f"{part_code}*.xlsx"]
            for pattern in glob_patterns:
                for candidate_path in sorted(search_dir.glob(pattern)):
                    stem_upper = candidate_path.stem.upper()
                    if mold_code and mold_code.upper() not in stem_upper:
                        continue
                    return candidate_path

        if explicit_excel_path:
            stem = explicit_excel_path.stem.upper()
            if normalized_part_code in stem:
                return explicit_excel_path

        return None

    def _parse_nc_excel_workbook(self, workbook_path: Path) -> Dict[str, Any]:
        """旧版 Excel workbook 解析逻辑（已停用，保留作参考，不删除）"""
        workbook = load_workbook(filename=str(workbook_path), data_only=True, read_only=True)
        try:
            nc_details = []

            for sheet_name, face_code in self.EXCEL_FACE_CODE_MAP.items():
                details = self._parse_nc_excel_sheet(workbook, sheet_name)
                nc_details.append({
                    "face_code": face_code,
                    "details": details
                })

            return {"nc_details": nc_details}
        finally:
            workbook.close()

    def _parse_nc_excel_sheet(self, workbook, sheet_name: str) -> List[Dict[str, Any]]:
        """旧版 Excel sheet 解析逻辑（已停用，保留作参考，不删除）"""
        if sheet_name not in workbook.sheetnames:
            self.logger.warning(f"[NCTimeAgent] Excel 缺少工作表: {sheet_name}")
            return []

        worksheet = workbook[sheet_name]
        header_row_index = None
        column_indexes: Dict[str, int] = {}

        for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            normalized_values = [str(value).strip() if value is not None else "" for value in row]
            if "程序名称" in normalized_values and "时间" in normalized_values:
                header_row_index = row_index
                column_indexes = {
                    "program_name": normalized_values.index("程序名称"),
                    "time_value": normalized_values.index("时间"),
                    "tool_name": normalized_values.index("刀具") if "刀具" in normalized_values else -1,
                }
                break

        if not header_row_index:
            self.logger.warning(f"[NCTimeAgent] Excel 工作表未找到 NC 表头: {sheet_name}")
            return []

        details: List[Dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            program_name = self._get_excel_cell_value(row, column_indexes.get("program_name"))
            time_value = self._to_decimal(self._get_excel_cell_value(row, column_indexes.get("time_value")))
            tool_name = self._get_excel_cell_value(row, column_indexes.get("tool_name"))

            if not program_name or time_value is None:
                continue

            code = self._extract_excel_operation_code(program_name, tool_name)
            if not code:
                self.logger.warning(f"[NCTimeAgent] Excel 程序名无法提取 code: sheet={sheet_name}, program={program_name}")
                continue

            details.append({
                "code": code,
                "value": time_value,
                "program_name": program_name,
                "tool": tool_name or ""
            })

        return details

    def _get_excel_cell_value(self, row: tuple, index: Optional[int]) -> str:
        if index is None or index < 0 or index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        return str(value).strip()

    def _to_decimal(self, value: Any) -> Optional[Decimal]:
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value).strip())
        except Exception:
            return None

    def _extract_excel_operation_code(self, program_name: str, tool_name: str = "") -> str:
        operation_name = (program_name or "").strip()
        if not operation_name:
            return ""

        chinese_code = ""
        if "开粗" in operation_name:
            chinese_code = "开粗"
        elif "半精" in operation_name:
            chinese_code = "半精"
        elif "全精" in operation_name:
            chinese_code = "全精"
        elif "台阶线割槽" in operation_name:
            chinese_code = "台阶线割槽"
        elif "线割槽" in operation_name:
            chinese_code = "线割槽"
        elif "预钻" in operation_name:
            chinese_code = "预钻"
        elif "精" in operation_name:
            chinese_code = "精铣"

        if chinese_code:
            return chinese_code

        parts = [part.strip() for part in operation_name.split("_") if part and part.strip()]
        if len(parts) >= 3 and parts[-1] == "台阶":
            return parts[-3]
        if len(parts) >= 2:
            return parts[-2]

        tool_value = (tool_name or "").strip()
        return tool_value or operation_name

    def _extract_json_face_code(self, operation: Dict[str, Any]) -> str:
        face_letter = ""
        for param in operation.get("parameters", []):
            if str(param.get("id")) == "parent_group":
                value = str(param.get("value") or "").strip()
                match = re.match(r"^([A-F])", value, re.IGNORECASE)
                if match:
                    face_letter = match.group(1).upper()
                    break

        if not face_letter:
            operation_name = str(operation.get("operation_name") or "").strip()
            prefix = operation_name.split("_")[0].strip().upper() if "_" in operation_name else ""
            if re.fullmatch(r"[A-F]", prefix):
                face_letter = prefix

        return self.JSON_FACE_CODE_MAP.get(face_letter, "")

    def _extract_parameter_decimal(self, operation: Dict[str, Any], parameter_id: Any) -> Optional[Decimal]:
        target_id = str(parameter_id)
        for param in operation.get("parameters", []):
            if str(param.get("id")) != target_id:
                continue
            value = param.get("value")
            if value in (None, ""):
                return None
            try:
                return Decimal(str(value))
            except Exception:
                return None
        return None

    def _extract_json_tool_name(self, operation_name: str) -> str:
        text = (operation_name or "").strip()
        if not text:
            return ""

        parts = [part.strip() for part in text.split("_") if part and part.strip()]
        if len(parts) >= 3 and re.fullmatch(r"[A-F]", parts[0], re.IGNORECASE):
            return parts[-1]
        if len(parts) >= 2:
            return parts[-1]
        return ""

    def _extract_fail_itemcodes(self, nc_result: Dict[str, Any]) -> List[str]:
        raw_codes = nc_result.get("fail_itemcode")
        if raw_codes is None:
            raw_codes = nc_result.get("data", {}).get("fail_itemcode", [])
        if not isinstance(raw_codes, list):
            return []

        normalized_codes: List[str] = []
        seen = set()
        for code in raw_codes:
            text = self._extract_subgraph_id(str(code).strip())
            if not text or text in seen:
                continue
            seen.add(text)
            normalized_codes.append(text)
        return normalized_codes

    async def _save_nc_failed_itemcodes(self, job_id: str, fail_itemcodes: List[str]):
        async for db in get_db():
            result = await db.execute(
                select(Job.meta_data).where(Job.job_id == job_id)
            )
            current_meta = result.scalar_one_or_none() or {}
            if not isinstance(current_meta, dict):
                current_meta = {}

            updated_meta = dict(current_meta)
            updated_meta["nc_failed_itemcodes"] = fail_itemcodes
            updated_meta["nc_failed_updated_at"] = datetime.utcnow().isoformat()

            await db.execute(
                update(Job)
                .where(Job.job_id == job_id)
                .values(meta_data=updated_meta)
            )
            await db.commit()

            self.logger.info(
                f"[NCTimeAgent] 保存 NC 识别失败物料编码: job_id={job_id}, count={len(fail_itemcodes)}"
            )
            break
